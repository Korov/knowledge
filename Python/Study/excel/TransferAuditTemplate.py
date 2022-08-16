import os
import re

import openpyxl


def transfer_en():
    originFile = 'C:\\Users\\korov\\Desktop\\temp\\SIEM操作审计cn&en.xlsx'
    workbook = openpyxl.load_workbook(originFile)
    worksheet = workbook["工作表1"]

    # columnIndex = 49
    # cellValue = worksheet.cell(columnIndex, 1).value
    # originTemplate = str(cellValue)
    # transferTemplate = originTemplate
    # matchResult = re.findall(r'(?<=[ "])\$\w+(?=[ ":.])', originTemplate, re.UNICODE | re.MULTILINE)
    # for resultIndex, result in enumerate(matchResult):
    #     transferTemplate = transferTemplate.replace(result, "{" + str(resultIndex) + "}", 1)
    # print(f"index:{columnIndex}, transfer: {transferTemplate}")

    keyIndex = 1
    for columnIndex in range(2, worksheet.max_row):
        cellValue = worksheet.cell(columnIndex, 2).value
        if cellValue is None:
            # print(f"index:{columnIndex} is none continue")
            continue
        originTemplate = str(cellValue)
        origin = originTemplate
        # en
        originTemplate = re.sub(r'((?<=(: ))|(?<=(:  )))[\u4e00-\u9fa5]+[段包]1.*', '$multipleValue', originTemplate)
        originTemplate = re.sub(r'(("\$?(name|group)\d")(,|、)?){3}（...）', '$multipleSplitValue', originTemplate)

        # zh
        # originTemplate = re.sub(r'((?<=(：))|(?<=(:  )))[\u4e00-\u9fa5]+[段包]1.*', '$multipleValue', originTemplate)
        # originTemplate = re.sub(r'((“\$?(name|group|tag)\d”)(,|、|，)?){3}（...）', '$multipleSplitValue', originTemplate)

        originTemplate = re.sub(r'((\$?(name|tag)\d)、?){3}（...）', '$multipleSplitValue', originTemplate)
        transferTemplate = originTemplate

        # en
        matchResult = re.findall(r'((?<=[ ">/])\$\w+((?=[-/ ":.$])|($)))|(【[\w/ ]*】)', originTemplate,
                                 re.UNICODE | re.MULTILINE)

        # zh
        # matchResult = re.findall(r'(\$\w+((?=[\u4e00-\u9fa5-/ ，、”:。$])|($)))|(【.*】)', originTemplate, re.UNICODE | re.MULTILINE)
        for index, results in enumerate(matchResult):
            for result in results:
                if result != '':
                    transferTemplate = transferTemplate.replace(result, "{" + str(index) + "}", 1)
        str(keyIndex).zfill(5)
        # print(f"index:{columnIndex}, origin: {origin}")
        # print(f"index:{columnIndex}, origin: {originTemplate}")
        # print(f"index:{columnIndex}, transfer: {transferTemplate}")
        # print(f"index:{str(keyIndex).zfill(5)}, origin: {origin}, transfer: {transferTemplate}")
        print(f"{str(keyIndex).zfill(5)} = {transferTemplate}")
        keyIndex = keyIndex + 1


def transfer_zh():
    originFile = 'C:\\Users\\korov\\Desktop\\temp\\SIEM操作审计cn&en.xlsx'
    workbook = openpyxl.load_workbook(originFile)
    worksheet = workbook["工作表1"]

    # columnIndex = 49
    # cellValue = worksheet.cell(columnIndex, 1).value
    # originTemplate = str(cellValue)
    # transferTemplate = originTemplate
    # matchResult = re.findall(r'(?<=[ "])\$\w+(?=[ ":.])', originTemplate, re.UNICODE | re.MULTILINE)
    # for resultIndex, result in enumerate(matchResult):
    #     transferTemplate = transferTemplate.replace(result, "{" + str(resultIndex) + "}", 1)
    # print(f"index:{columnIndex}, transfer: {transferTemplate}")

    keyIndex = 1
    for columnIndex in range(2, worksheet.max_row):
        cellValue = worksheet.cell(columnIndex, 1).value
        if cellValue is None:
            # print(f"index:{columnIndex} is none continue")
            continue
        originTemplate = str(cellValue)
        origin = originTemplate

        # zh
        originTemplate = re.sub(r'(((?<=(：))|(?<=(:  )))[\u4e00-\u9fa5]+[段包]1.*)|((字段\d“\$field\d”、?){3})',
                                '$multipleValue', originTemplate)
        originTemplate = re.sub(r'((“\$?(name|group|tag|value)\d”)(,|、|，)?){3}（...）', '$multipleSplitValue',
                                originTemplate)

        originTemplate = re.sub(r'((\$?(name|tag)\d)、?){3}（...）', '$multipleSplitValue', originTemplate)
        transferTemplate = originTemplate

        # zh
        matchResult = re.findall(r'(\$\w+((?=[\u4e00-\u9fa5-/ ，、”:。$])|($)))|(【.*】)', originTemplate,
                                 re.UNICODE | re.MULTILINE)
        for index, results in enumerate(matchResult):
            for result in results:
                if result != '':
                    transferTemplate = transferTemplate.replace(result, "{" + str(index) + "}", 1)
        str(keyIndex).zfill(5)
        # print(f"index:{columnIndex}, origin: {origin}")
        # print(f"index:{columnIndex}, origin: {originTemplate}")
        print(f"index:{columnIndex}, transfer: {transferTemplate}")
        # print(f"index:{str(keyIndex).zfill(5)}, origin: {origin}, transfer: {transferTemplate}")
        # print(f"{str(keyIndex).zfill(5)} = {transferTemplate}")
        keyIndex = keyIndex + 1


def generate_audi_template():
    origin_file = 'C:\\Users\\korov\\Desktop\\temp\\SIEM操作审计cn&en.xlsx'
    target_file = 'C:\\Users\\korov\\Desktop\\temp\\SIEM操作审计cn&en_bak.xlsx'
    if os.path.exists(target_file):
        os.remove(target_file)
    workbook = openpyxl.load_workbook(origin_file)
    worksheet = workbook["工作表1"]

    for column_index in range(2, worksheet.max_row):
        origin_cell_zh = worksheet.cell(column_index, 1).value
        origin_cell_en = worksheet.cell(column_index, 2).value
        if origin_cell_zh is None or origin_cell_en is None:
            continue

        origin_value_zh = str(origin_cell_zh)
        origin_value_en = str(origin_cell_en)

        zh_match_result = re.findall(r'【[\u4e00-\u9fa5/_\-\w ]*】', origin_value_zh,
                                     re.UNICODE | re.MULTILINE)
        en_match_result = re.findall(r'【[\u4e00-\u9fa5/_\-\w ]*】', origin_value_en,
                                     re.UNICODE | re.MULTILINE)
        if len(zh_match_result) != len(en_match_result):
            print(f'column index:{column_index}, match count not match, zh:{zh_match_result}, en:{en_match_result}')
        else:
            worksheet.cell(column_index, 5).value = str(zh_match_result)
            worksheet.cell(column_index, 6).value = str(en_match_result)
            for match_index in range(0, len(zh_match_result)):
                origin_value_zh = origin_value_zh.replace(zh_match_result[match_index], f'$fixed_value{match_index}')
                origin_value_en = origin_value_en.replace(en_match_result[match_index], f'$fixed_value{match_index}')

        zh_match_result = re.findall(
            r'((?<=(：))[\u4e00-\u9fa5]+[段包]1.*)|((字段\d“\$field\d”、?){3})|((“\$(group|name|tag|value)\d”，?){3}（...）)',
            origin_value_zh, re.UNICODE)
        en_match_result = re.findall(
            r'((?<=(: ))[\u4e00-\u9fa5]+[段包]1.*)|(("\$(group|name|tag|value)\d(")?[,、]?){3}（...）)',
            origin_value_en, re.UNICODE)

        valid_zh_match = []
        valid_en_match = []
        if len(zh_match_result) >= 1:
            if zh_match_result[0][0] != '':
                valid_zh_match.append(zh_match_result[0][0])
            if zh_match_result[0][2] != '':
                valid_zh_match.append(zh_match_result[0][2])
            if zh_match_result[0][4] != '':
                valid_zh_match.append(zh_match_result[0][4])

        if len(en_match_result) >= 1:
            if en_match_result[0][0] != '':
                valid_en_match.append(en_match_result[0][0])
            if en_match_result[0][2] != '':
                valid_en_match.append(en_match_result[0][2])

        if len(valid_zh_match) != len(valid_en_match):
            print(
                f'column index:{column_index}, match count not match, zh:{valid_zh_match}, en:{valid_en_match}, origin zh:{origin_value_zh}, origin en:{origin_value_en}')
        else:
            if len(valid_zh_match) == 1:
                origin_value_zh = origin_value_zh.replace(valid_zh_match[0], f'$dynamic_value0')
                origin_value_en = origin_value_en.replace(valid_en_match[0], f'$dynamic_value0')
                worksheet.cell(column_index, 7).value = valid_zh_match[0]
                worksheet.cell(column_index, 8).value = valid_en_match[0]

        zh_match_result = re.findall(r'\$[\w]+', origin_value_zh, re.ASCII)
        en_match_result = re.findall(r'\$[\w]+', origin_value_en, re.ASCII)

        print(f'zh keys:{zh_match_result}, en keys:{en_match_result}')
        if len(zh_match_result) != len(en_match_result):
            print(f'column index:{column_index}, not match count, zh match:{zh_match_result}, en match:{en_match_result}')
        else:
            zh_match_map = {}
            for value_index, result_value in enumerate(zh_match_result):
                zh_match_map.setdefault(result_value, value_index)
            en_match_map = {}
            for result_value in en_match_result:
                if zh_match_map.get(result_value) is not None:
                    en_match_map.setdefault(result_value, zh_match_map.get(result_value))

                    origin_value_zh = origin_value_zh.replace(result_value,
                                                              '{' + str(zh_match_map.get(result_value)) + '}')
                    origin_value_en = origin_value_en.replace(result_value,
                                                              '{' + str(zh_match_map.get(result_value)) + '}')
                # else:
                    # print(f'column index:{column_index}, not match key:{result_value} zh match:{zh_match_result}, en match:{en_match_result}')

        # print(f"{str(column_index - 1).zfill(5)} = {origin_value_zh}")
        # worksheet.cell(column_index, 3).value = origin_value_zh
        # worksheet.cell(column_index, 4).value = origin_value_en
        # workbook.save('C:\\Users\\korov\\Desktop\\temp\\SIEM操作审计cn&en_bak.xlsx')
        workbook.close()


if __name__ == "__main__":
    # transfer_en()
    # transfer_zh()
    generate_audi_template()
    # message = "用户：$user 从IP：$clientip 编辑自定义分组“$name”：编辑字段1“$mutiValue1”，编辑字段2“$mutiValue2”，编辑字段3“$mutiValue3”(...)"
    # zh_match_result = re.match(
    #     r'((?<=(：))[\u4e00-\u9fa5]+[段包]1[“”\$\w_\-1]+)|((字段\d“\$field\d”、?){3})|((“\$(group|name|tag|value)\d”，?){3}（...）)',
    #     message, re.UNICODE)
    # print(f'result:{zh_match_result}')
