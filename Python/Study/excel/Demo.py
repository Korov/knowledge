import openpyxl
import datetime
import time

from openpyxl.worksheet import worksheet

if __name__ == "__main__":
    originaFile = 'C:\\Users\\Korov\\Desktop\\temp\\demo.xlsx'
    targetFile = 'C:\\Users\\Korov\\Desktop\\temp\\demo1.xlsx'
    workbook = openpyxl.load_workbook(originaFile)
    worksheet = workbook["Sheet0"]
    worksheet.max_row
    print(worksheet.max_row)
    print(str(worksheet.cell(278, 1).value))
    print(worksheet.cell(278, 1).comment)
    fill = openpyxl.styles.PatternFill("solid", fgColor="1874CD")
    worksheet.cell(278, 1).fill = fill
    workbook.save(targetFile)