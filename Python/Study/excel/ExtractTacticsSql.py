import openpyxl

if __name__ == "__main__":
    origin_file_zh = 'C:\\Users\\korov\\Downloads\\MITRE ATT&CK   矩阵技术描述列表（中文版)_20220919.xlsx'
    workbook_zh = openpyxl.load_workbook(origin_file_zh)

    origin_file_en = 'C:\\Users\\korov\\Downloads\\MITRE ATT&CK   矩阵技术描述列表（英文版)_20221123.xlsx'
    workbook_en = openpyxl.load_workbook(origin_file_en)

    techniques_map = dict()
    update_sql_format = "UPDATE `siem_att_ck_info` SET `name_en` = '{}', `description_en` = '{}' WHERE `code` = '{}' and `parent_code` = '{}';"
    for worksheet_index in range(0, len(workbook_zh.worksheets)):
        worksheet = workbook_zh.worksheets[worksheet_index]
        worksheet_en = workbook_en.worksheets[worksheet_index]
        merge_cells = []
        max_row = 0
        for cell in worksheet.merged_cells.ranges:
            if cell.max_col == 4:
                if cell.min_row >= 3:
                    merge_cells.append((cell.min_col, cell.min_row, cell.max_row))
            elif cell.max_col == 1:
                max_row = cell.max_row

        merge_cells.sort()

        cells = []
        sqls = []
        update_sqls = []
        for cell_index in range(0, len(merge_cells)):
            current_cell = merge_cells[cell_index]
            current_col = current_cell[0]
            current_min_row = current_cell[1]
            current_max_row = current_cell[2]
            if cell_index == 0 and current_min_row > 3:
                for row in range(3, current_min_row):
                    cells.append((current_col, row, row))

            cells.append(current_cell)

            if cell_index == len(merge_cells) - 1:
                if current_max_row < max_row:
                    for row in range(current_max_row + 1, max_row + 1):
                        cells.append((current_col, row, row))

            if cell_index < len(merge_cells) - 1:
                next_cell = merge_cells[cell_index + 1]
                next_col = next_cell[0]
                next_min_row = next_cell[1]
                next_max_row = next_cell[2]

                if current_max_row + 1 != next_min_row:
                    for row in range(current_max_row + 1, next_min_row):
                        cells.append((current_col, row, row))

        cells.sort()
        tactics_code = worksheet.cell(column=1, row=3).value.strip()
        tactics_name = worksheet.cell(column=2, row=3).value.strip()
        tactics_description = worksheet.cell(column=3, row=3).value.strip().replace("\\", "\\\\")

        tactics_code_en = worksheet_en.cell(column=1, row=3).value.strip()
        tactics_name_en = worksheet_en.cell(column=2, row=3).value.strip()
        tactics_description_en = worksheet_en.cell(column=3, row=3).value.strip().replace("\\", "\\\\").replace("'",
                                                                                                                "\\'")

        if tactics_code != tactics_code_en:
            print(
                f"work sheet:{worksheet_index}, colum:{1}, row:{3}, tactics code:{tactics_code}, tactics code en:{tactics_code_en} error")
            exit(-1)

        sqls.append(
            'INSERT INTO `siem_att_ck_info`(`code`, `parent_code`, `type`, `name_zh`, `description_zh`, `name_en`, `description_en`, `create_time`, `update_time`) VALUES')
        sqls.append(
            "('{}', '', 0, '{}', '{}', '', '', NOW(), NOW()),".format(tactics_code, tactics_name,
                                                                      tactics_description))
        update_sqls.append(update_sql_format.format(tactics_name_en, tactics_description_en, tactics_code, ""))

        for cell_index in range(0, len(cells)):
            cell = cells[cell_index]
            col = cell[0]
            min_row = cell[1]
            max_row = cell[2]

            techniques_code = worksheet.cell(column=col, row=min_row).value.strip()
            techniques_name = worksheet.cell(column=col + 1, row=min_row).value.strip()
            techniques_description = worksheet.cell(column=col + 2, row=min_row).value.strip().replace("\\", "\\\\")

            techniques_code_en = worksheet_en.cell(column=col, row=min_row).value.strip()
            techniques_name_en = worksheet_en.cell(column=col + 1, row=min_row).value.strip()
            techniques_description_en = worksheet_en.cell(column=col + 2, row=min_row).value.strip().replace("\\",
                                                                                                             "\\\\").replace(
                "'", "\\'")

            if techniques_code != techniques_code_en:
                print(
                    f"work sheet:{worksheet_index}, colum:{col}, row:{min_row}, techniques_code:{techniques_code}, techniques_code_en:{techniques_code_en} error")
                exit(-1)

            if techniques_map.keys().__contains__(techniques_code):
                techniques_map[techniques_code].append(tactics_code)
                sqls.append(
                    "('{}', '{}', 1, '{}', '{}', '', '', NOW(), NOW()),".format(techniques_code, tactics_code,
                                                                                techniques_name,
                                                                                techniques_description
                                                                                ))
                update_sqls.append(
                    update_sql_format.format(techniques_name_en, techniques_description_en, techniques_code,
                                             tactics_code))

                continue
            else:
                techniques_map[techniques_code] = [tactics_code]
                sqls.append(
                    "('{}', '{}', 1, '{}', '{}', '', '', NOW(), NOW()),".format(techniques_code, tactics_code,
                                                                                techniques_name,
                                                                                techniques_description))
                update_sqls.append(
                    update_sql_format.format(techniques_name_en, techniques_description_en, techniques_code,
                                             tactics_code))

            for row_index in range(min_row, max_row + 1):
                sub_techniques_cell = worksheet.cell(column=col + 3, row=row_index)
                if sub_techniques_cell.value == None:
                    continue
                sub_techniques_code = worksheet.cell(column=col + 3, row=row_index).value.strip()
                sub_techniques_name = worksheet.cell(column=col + 4, row=row_index).value.strip()
                sub_techniques_description = worksheet.cell(column=col + 5, row=row_index).value.strip().replace("\\",
                                                                                                                 "\\\\")

                sub_techniques_code_en = worksheet_en.cell(column=col + 3, row=row_index).value.strip()
                sub_techniques_name_en = worksheet_en.cell(column=col + 4, row=row_index).value.strip()
                sub_techniques_description_en = worksheet_en.cell(column=col + 5, row=row_index).value.strip().replace(
                    "\\", "\\\\").replace("'", "\\'")

                if sub_techniques_code != sub_techniques_code_en:
                    print(
                        f"work sheet:{worksheet_index}, colum:{col + 3}, row:{row_index}, sub_techniques_code:{sub_techniques_code}, sub_techniques_code_en:{sub_techniques_code_en} error")
                    exit(-1)
                sqls.append("('{}', '{}', 2, '{}', '{}', '', '', NOW(), NOW()),".format(sub_techniques_code,
                                                                                        techniques_code,
                                                                                        sub_techniques_name,
                                                                                        sub_techniques_description))
                update_sqls.append(
                    update_sql_format.format(sub_techniques_name_en, sub_techniques_description_en, sub_techniques_code,
                                             techniques_code))

        # for sql_index in range(0, len(sqls)):
        #     sql = sqls[sql_index]
        #     if sql_index == len(sqls) - 1:
        #         sql = sql[:len(sql) - 1] + ";"
        #     print(sql)

        for sql_index in range(0, len(update_sqls)):
            sql = update_sqls[sql_index]
            print(sql)
        print("\n\n")

    print("finish")

    # print('workbook title:{}, max column:{}, max row:{}'.format(worksheet.title, worksheet.max_column,
    #                                                             worksheet.max_row))
