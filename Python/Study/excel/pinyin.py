import openpyxl
import xpinyin
from loguru import logger

logger.add('insert1.log', rotation="200 MB",
           format="{message}",
           level="INFO")

# Sheet1
if __name__ == "__main__":
    origin_file_zh = 'D:\\temp\\Book2.xlsx'
    workbook_zh = openpyxl.load_workbook(origin_file_zh)
    worksheet = workbook_zh.worksheets[0]
    pinyin = xpinyin.Pinyin()
    for columnIndex in range(1, worksheet.max_row):
        cellValue = worksheet.cell(columnIndex, 1).value
        pinyinValue = worksheet.cell(columnIndex, 2).value
        pinyinValue = pinyinValue + str(columnIndex)
        logger.info(f"('{pinyinValue}', '{cellValue}', 1, '2023-02-15 10:44:26', '2023-02-15 10:44:26'),")

    print("debug")