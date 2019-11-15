import datetime
import time

import openpyxl

# 写excel
from util import StringUtils


def write_excel():
    # 创建工作簿
    workBook = openpyxl.Workbook()

    workBook.create_sheet("my sheet1", 3)
    workBook.create_sheet("my sheet2", 2)
    workBook.create_sheet("my sheet3", 1)

    # 获取第一个sheet
    workSheet = workBook.active

    # 在特定单元格写入数据
    workSheet["A1"] = 42
    workSheet["B1"] = "你好" + "hello"

    # 写入多个单元格
    workSheet.append([1, 2, 3])

    workSheet["A3"] = datetime.datetime.now()  # 写入一个当前时间
    # 写入一个自定义的时间格式
    workSheet["A4"] = time.strftime("%Y{y}%m{m}%d{d} %H{h}%M{f}%S{s}").format(y="年", m="月", d="日", h="时", f="分", s="秒")

    workSheet = workBook['my sheet2']
    workSheet.merge_cells(range_string='A1:E3', start_row=1, start_column=1, end_row=3, end_column=5)
    workSheet.merge_cells(range_string='B5:F7')
    workSheet.merge_cells(start_row=9, start_column=1, end_row=11, end_column=5)

    workSheet.merge_cells(range_string='A15:E18')
    workSheet.unmerge_cells(range_string='A15:E18')

    currentTime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    savePath = "C:\\Users\\Korov\\Desktop\\temporary\\PythonExcel\\sample{}.xlsx".format(currentTime)
    # Save the file
    workBook.save(savePath)


"""
只可以处理xlsx文件
可以读取，写入和修改xlsx文件
处理功能复杂的文件
可以处理数据量很大的文件
可以跨平台
"""


def checkLength(workSheet, row, cols, param):
    maxNum = 0
    for col in range(1, cols + 1):
        cellValue = StringUtils.remove_line_breaks(str(workSheet.cell(row=row, column=col).value))
        if cellValue.find(param) >= 0:
            cells = cellValue.split(param)
            if maxNum == 0:
                maxNum = len(cells)
            elif maxNum != len(cells):
                print('第{}行，第{}列，分号与其他列数量不一样'.format(row, col))
    print('第{}行，maxNum:{}'.format(row, maxNum))
    return maxNum


def addRange(maxNum, row, cols, workSheet, param):
    workSheet.insert_rows(row + 1, maxNum - 1)
    for col in range(1, cols + 1):
        cellValue = StringUtils.remove_line_breaks(str(workSheet.cell(row=row, column=col).value))
        for newRow in range(row, row + maxNum):
            cellValueLength = cellValue.find(param) if cellValue.find(param) >= 0 else len(cellValue)
            cellValueTemp = cellValue[0:cellValueLength]
            workSheet.cell(row=newRow, column=col, value=cellValueTemp)
            cellValue = cellValue[cellValue.find(param) + 1:]


def sheetCopy(originalSheet, targeSheet):
    for row in range(1, originalSheet.max_row + 1):
        for col in range(1, originalSheet.max_column + 1):
            targeSheet.cell(row=row, column=col,
                            value=StringUtils.remove_line_breaks(str(originalSheet.cell(row=row, column=col).value)))


def unmergeSheet(sheet):
    mergedCells = sheet.merged_cells

    while len(mergedCells.ranges) > 0:
        mergedCell = mergedCells.ranges[0]
        cell = sheet.cell(mergedCell.min_row, mergedCell.min_col)
        sheet.unmerge_cells(range_string=mergedCell.coord)
        for row in range(mergedCell.min_row, mergedCell.max_row + 1):
            for col in range(mergedCell.min_col, mergedCell.max_col + 1):
                sheet.cell(row=row, column=col, value=StringUtils.remove_line_breaks(str(cell.value)))


def splitSheetValue(sheet, strSplit):
    row = 0
    while row < sheet.max_row:
        row = row + 1
        maxNum = checkLength(sheet, row, sheet.max_column, strSplit)
        if maxNum == 0:
            continue
        addRange(maxNum, row, sheet.max_column, sheet, strSplit)


def getWorkBook(originaFile):
    return openpyxl.load_workbook(originaFile)


def getWorkSheet(workbook=None, filename=None, sheetname=None):
    if filename is not None:
        workbook = getWorkBook(filename)
    return workbook[sheetname]


def format_excel(originaFile, targetFile, strSplit, sheetName, newSheetName):
    workBook = getWorkBook(originaFile)
    workSheet = getWorkSheet(workbook=workBook, sheetname=sheetName)
    unmergeSheet(workSheet)
    newWorkSheet = workBook.create_sheet(newSheetName)
    sheetCopy(workSheet, newWorkSheet)
    splitSheetValue(newWorkSheet, strSplit)
    # Save the file
    workBook.save(targetFile)


def readExcelToList(workSheet):
    dataList = []
    for row in range(1, workSheet.max_row + 1):
        dataRow = []
        for col in range(1, workSheet.max_column + 1):
            dataRow.append(workSheet.cell(row=row, column=col).value)
        dataList.append(dataRow)
    return dataList


def getAttrbuites(dataList):
    return "(`{}`)".format(str.join("`, `", dataList[0]))


def getValue(row, dataList):
    return "('{}')".format(str.join("', '", dataList[row]))


def getValues(dataList):
    lineValues = []
    for row in range(1, len(dataList)):
        lineValues.append(getValue(row, dataList))
    return str.join(",\n", lineValues)


def writeFile(sql, sqlFile):
    file_handle = open(sqlFile, mode='w')
    file_handle.write(sql)
    file_handle.close()


def joinTheValue(tableName, attributies, values):
    return "insert into `{}` {} values\n{};".format(tableName, attributies, values)


def exportToSql(originaFile, sheetName, sqlFile):
    workSheet = getWorkSheet(filename=originaFile, sheetname=sheetName)
    dataList = readExcelToList(workSheet)
    tableName = "aaa"
    attributies = getAttrbuites(dataList)
    values = getValues(dataList)
    sql = joinTheValue(tableName, attributies, values)
    writeFile(sql, sqlFile)


if __name__ == "__main__":
    originaFile = 'C:\\Users\\Korov\\Desktop\\temporary\\PythonExcel\\test.xlsx'
    currentTime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    targetFile = "C:\\Users\\Korov\\Desktop\\temporary\\PythonExcel\\sample{}.xlsx".format(currentTime)
    strSplit = ';'
    sheetName = "test"
    newSheetName = "newTest"
    # 写入Excel
    # write_excel();
    # format_excel(originaFile, targetFile, strSplit, sheetName, newSheetName);

    originaFile = 'C:\\Users\\Korov\\Desktop\\temporary\\PythonExcel\\sample.xlsx'
    sheetName = "newTest"
    sqlFile = 'C:\\Users\\Korov\\Desktop\\temporary\\PythonExcel\\sample.sql'
    exportToSql(originaFile, sheetName, sqlFile)
    print("写入成功")
